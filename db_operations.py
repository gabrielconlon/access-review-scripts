import sqlite3
import json
import logging
from datetime import datetime
import openpyxl

# Load configuration
with open('config.json', 'r') as config_file, open('config.json', 'r') as config_file:
    config = json.load(config_file)

# Configure logging
logging.basicConfig(filename=config['log_file'], level=getattr(logging, config['log_level']), format=config['log_format'])

class Service:
    def __init__(self, service_name, fields):
        self.service_name = service_name
        self.fields = fields

class User:
    def __init__(self, display_name, mail=None, created_from=None):
        self.display_name = display_name
        self.mail = mail
        self.created_from = created_from
        self.services = {}

    def add_service(self, service):
        self.services[service.service_name] = service.fields

def create_service_table(service_name, headers, db_conn):
    cursor = db_conn.cursor()
    columns = ", ".join([f'"{header}" TEXT' for header in headers])
    cursor.execute(f'''CREATE TABLE IF NOT EXISTS "{service_name}" ({columns})''')
    db_conn.commit()

    # Check for missing columns and add them
    cursor.execute(f'PRAGMA table_info("{service_name}")')
    existing_columns = {row[1] for row in cursor.fetchall()}
    for header in headers:
        if header not in existing_columns:
            cursor.execute(f'ALTER TABLE "{service_name}" ADD COLUMN "{header}" TEXT')
    db_conn.commit()
    print(f"Table '{service_name}' created or updated with necessary columns.")

def insert_service_data(service_name, user_data, db_conn):
    cursor = db_conn.cursor()
    columns = ", ".join([f'"{key}"' for key in user_data.keys()])
    placeholders = ", ".join(["?" for _ in user_data])
    values = [str(value) for value in user_data.values()]  # Convert all values to strings
    cursor.execute(f'''INSERT OR REPLACE INTO "{service_name}" ({columns}) VALUES ({placeholders})''', values)
    db_conn.commit()
    print(f"Inserted/Updated data in table '{service_name}'.")

def convert_datetime_to_str(data):
    if isinstance(data, dict):
        return {key: convert_datetime_to_str(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [convert_datetime_to_str(item) for item in data]
    elif isinstance(data, datetime):
        return data.isoformat()
    else:
        return data

def process_workbook(file_path, db_conn, verbosity):
    print(f"Processing workbook: {file_path}")
    workbook = openpyxl.load_workbook(file_path)
    users = {}

    # Process all sheets except excluded ones
    for sheet_name in workbook.sheetnames:
        if sheet_name.lower() in config['excluded_sheets']:
            continue

        if verbosity > 0:
            print(f"Processing sheet: {sheet_name}")
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        headers = [header if header is not None else f"Column_{i}" for i, header in enumerate(headers)]
        if verbosity > 1:
            print(f"Headers: {headers}")

        # Create or update the service table
        create_service_table(sheet_name, headers, db_conn)

        for row in sheet.iter_rows(min_row=2, values_only=True):
            user_data = dict(zip(headers, row))
            login = None
            display_name = None
            for header, cell in zip(headers, row):
                if header.lower() in config['login_columns'] and cell is not None:
                    login = cell.lower()  # Convert login to lowercase for case-insensitive comparison
                    break
                elif header.lower() == "displayname":
                    display_name = cell

            if login or display_name:
                user_key = login if login else display_name
                if user_key not in users:
                    users[user_key] = User(display_name, login, sheet_name)
                users[user_key].add_service(Service(sheet_name, user_data))
                if verbosity > 1:
                    print(f"Added service for user: {user_key}")

            # Insert or update the service data in the service table
            insert_service_data(sheet_name, user_data, db_conn)

    # Insert users into SQLite
    cursor = db_conn.cursor()
    cursor.execute('''CREATE TABLE IF NOT EXISTS users (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        mail TEXT,
                        display_name TEXT,
                        services TEXT,
                        created_from TEXT
                      )''')
    print("Table 'users' created or already exists.")

    for key, user in users.items():
        services_str = json.dumps(convert_datetime_to_str(user.services))
        cursor.execute('''INSERT OR REPLACE INTO users (mail, display_name, services, created_from) VALUES (?, ?, ?, ?)''', (user.mail, user.display_name, services_str, user.created_from))
        if verbosity > 1:
            print(f"Inserted/Updated user: {user.mail} with display name: {user.display_name}")

    db_conn.commit()
    print("Database commit successful.")